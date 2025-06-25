#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from pathlib import Path
import importlib
import win32com.client
import warnings
import time
import os
from datetime import datetime
from openpyxl.styles import PatternFill
warnings.filterwarnings("ignore")  # Oculta todos los warnings

# Mis modulos
import download_mail_files as dmf
import file_management as fm
import join_dataframes as jdf
import load_data as load
import send_email as sm
import print_management as pm
import log_management as log


# In[157]:


# import pyfiglet
# print(pyfiglet.figlet_format("Plantilla Yape"))

banner = r"""
 ____  _             _   _ _ _        __   __               
|  _ \| | __ _ _ __ | |_(_) | | __ _  \ \ / /_ _ _ __   ___ 
| |_) | |/ _` | '_ \| __| | | |/ _` |  \ V / _` | '_ \ / _ \
|  __/| | (_| | | | | |_| | | | (_| |   | | (_| | |_) |  __/
|_|   |_|\__,_|_| |_|\__|_|_|_|\__,_|   |_|\__,_| .__/ \___|
                                                |_|         
        💥 AUTOMATIZADOR ENVIO DE YAPE PLANTILLA 💥
"""

print(banner, end='\n\n')
time.sleep(1)


# In[ ]:


# Configuracion

# Diccionarios
COLORS = { # Colores rellena celdas
    "CAM-": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    "PED-": PatternFill(start_color="99FF66", end_color="99FF66", fill_type="solid"),
    "CHA-": PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid"),
    "ATI-": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
}
outlook_folder_codes = {
    0: 'Calendario',
    1: 'Contactos',
    2: 'Borradores',
    3: 'Diario / Jornal',
    4: 'Notas',
    5: 'Tareas',
    6: 'Bandeja de entrada',
    7: 'Bandeja de salida',
    8: 'Elementos enviados',
    9: 'Elementos eliminados',
    10: 'Bandeja de correo del servidor',
    11: 'Conflictos',
    12: 'Elementos de sincronizacion local',
    13: 'Elementos de sincronizacion (Envio)',
    14: 'Elementos de sincronización (Recibo)',
    15: 'Elementos de sincronización completa',
    16: 'Diario de formularios',
    17: 'Carpeta de búsqueda',
    18: 'Bandeja para reglas cliente',
    19: 'Carpeta de sugerencias de correo',
}
outlook_object_types = {
    "AppointmentItem": 26,
    "MailItem": 43,
    "TaskItem": 46,
    "ContactItem": 48,
    "MeetingItem": 53,
}
MONTHS = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
    7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 
    11: 'Noviembre', 12: 'Diciembre'
}
CAMANA = {
    "name": "CAM-",
    "sender_email": "frans.castillo@aclogistica.pe",  # Nombre del asunto de correo
    "file_name": "",                                  # Nombre del archivo local
    "file_address": "",                               # Direccion del archivo local
    "mail_received_time": "",                         # Fecha del correo recibida
    "sheet_name": "Pedmap",
    "relevant_columns": [
        "CDTRA",
        "CHOFER",        
        "CDCLI",
        "NORSO",
    ]
}
PEDREGAL = {
    "name": "PED-",
    "sender_email": "wilson.contreras@aclogistica.pe",  # Nombre del asunto de correo
    "file_name": "",                                    # Nombre del archivo local
    "file_address": "",                                 # Direccion del archivo local
    "mail_received_time": "",                           # Fecha del correo recibida
    "sheet_name": "Pedmap",
    "relevant_columns": [
        "CDTRA",
        "CHOFER",        
        "CDCLI",
        "NORSO",
    ]
}
CHALA = {
    "name": "CHA-",
    "sender_email": "paolo.velasquez@aclogistica.pe",  # Nombre del asunto de correo
    "file_name": "",                                   # Nombre del archivo local
    "file_address": "",                                # Direccion del archivo local
    "mail_received_time": "",                          # Fecha del correo recibida
    "sheet_name": "Pedmap",
    "relevant_columns": [
        "CDTRA",
        "CHOFER",        
        "CDCLI",
        "NORSO",
    ]
}
ATICO = {
    "name": "ATI-",
    "sender_email": "admatico@ayacda.com",             # Nombre del asunto de correo
    "file_name": "",                                   # Nombre del archivo local
    "file_address": "",                                # Direccion del archivo local
    "mail_received_time": "",                          # Fecha del correo recibida
    "sheet_name": "CARGA",
    "relevant_columns": [
        "CDTRA",
        "Transportista 1",
        "CDCLI",
        "CLIENTE",
    ]
}
LOCACIONES = [CAMANA, PEDREGAL, CHALA, ATICO]
INVALID_CDTRA = ["8995", "8996", "8997", "", " ", "0", "N/A", "ND", None]

# Variables
MAPI = "MAPI" # Messaging Application Programming Interface
DOT = "."
OUTLOOK = "Outlook"
APPLICATION = "Application"
TODAY = datetime.today().strftime('%d-%m-%Y')
PROJECT_ADDRESS = r"C:\Users\AYACDA23\Desktop\PROG. YAPE\yape_plantilla"
FILE_TEMPLATE = os.path.join(PROJECT_ADDRESS, 'staticfiles/plantilla.xlsx')
PROCESSED_FILE = os.path.join(PROJECT_ADDRESS, f'Carga Plantilla AYA {TODAY}.xlsx')
SUBJECT = f'Carga Plantilla AYA {TODAY} // A Y A DISTRIBUCIONES EIRL'
BODY_MAIL = f'Estimados,\nAdjunto planificación de yape 25-06-2025.\nSaludos,'
FIRM = os.path.join(PROJECT_ADDRESS, 'staticfiles/firma.jpg')
MAIL_TO = "hugo.lino@prosegur.com;"
MAIL_CC = "juan.callan@prosegur.com;jorge.barboza@prosegur.com;marlon.solorzano@prosegur.com;contabilidad@ayacda.com;ADM@ayacda.com;asisconta@ayacda.com"
TEST_MAIL_TO = "ainformacion@ayacda.com;"
TEST_MAIL_CC = ";"


# ### Obtener correos de outlook
# 
# Descargar archivos de correo

# In[ ]:


importlib.reload(dmf)
importlib.reload(log)

mail_files = dmf.main(
    PROJECT_ADDRESS,
    LOCACIONES,
    OUTLOOK,
    DOT,
    APPLICATION,
    MAPI,
    outlook_object_types['MailItem']
)

log.delete_log(PROJECT_ADDRESS)


# Asignar configuraciones restantes

# In[ ]:


importlib.reload(pm)

def set_configurations(locacion):
    for clave, valores in mail_files.items():
        if locacion['name'] == clave:
            locacion['file_name'] = valores[0]
            locacion['file_address'] = valores[1]
            locacion['mail_received_time'] = valores[2]

    return locacion

CAMANA = set_configurations(CAMANA)
PEDREGAL = set_configurations(PEDREGAL)
CHALA = set_configurations(CHALA)
ATICO = set_configurations(ATICO)

for locacion in LOCACIONES:
    pm.show_document(locacion)


# ### Gestion y transformacion de datos

# In[ ]:


importlib.reload(fm)
importlib.reload(jdf)
importlib.reload(pm)

df_cam, df_ped, df_cha, df_ati = fm.main(CAMANA, PEDREGAL, CHALA, ATICO) # Obtener la data con las columnas filtradas
log.write_log(PROJECT_ADDRESS, '[*] Dataframes generados')
df_total = jdf.main(PROJECT_ADDRESS, CAMANA, PEDREGAL, CHALA, ATICO, df_cam, df_ped, df_cha, df_ati, INVALID_CDTRA) # Juntar los dataframes
pm.show_df(df_total)


# In[ ]:


importlib.reload(load)
importlib.reload(log)

load.main(FILE_TEMPLATE, df_total, PROCESSED_FILE, COLORS)
log.write_log(PROJECT_ADDRESS, '[*] Dataframe total cargado en excel')


# ### Envio de Correo

# In[ ]:


importlib.reload(sm)

log.read_log(PROJECT_ADDRESS, "log.txt")

# 🔁 Bucle de confirmación
while True:
    print("\nSelecciona opcion de correo:")
    print("  [1] Correos Oficiales")
    print("  [2] Correo de Prueba")
    print("  [3] Terminar Proceso")

    mail_option = int(input("\n>> Opción (1 - 3): "))
    if mail_option == 1:
        sm.main(OUTLOOK, DOT, APPLICATION, MAIL_TO, MAIL_CC, PROCESSED_FILE, SUBJECT, FIRM)
    elif mail_option == 2:
        sm.main(OUTLOOK, DOT, APPLICATION, TEST_MAIL_TO, TEST_MAIL_CC, PROCESSED_FILE, SUBJECT, FIRM)
    else:
        print("\n[✓] Proceso finalizado.\n")
        break  # salir del bucle


# ### Limpiar Carpeta

# In[ ]:


carpeta = Path(PROJECT_ADDRESS)
#print(list(carpeta.glob('*.png')))

# Eliminar archivos .png
for file in carpeta.glob('*.xlsx'):
    file.unlink()  # .unlink() elimina el archivo    


# ### Export it as .py
