# Modulo encargado de agregar datos creando un nuevo archivo mediante el uso de la plantilla.xlsx
from openpyxl import load_workbook

def main(
        FILE_TEMPLATE,
        df_total,
        PROCESSED_FILE,
        COLORS
):
    # Cargar archivo
    wb = load_workbook(FILE_TEMPLATE)
    ws = wb["CARGA"]  # Usa el nombre exacto de la hoja

    # Supongamos que tienes datos nuevos como lista de listas
    # nuevas_filas = [
    #     ["CAM-", 6014, "NUEVO TRANSPORTISTA", " ", " ", " ", " ", " ", 1234567, "NUEVO CLIENTE"],
    #     ["CAM-", 6010, "OTRO TRANSPORTISTA", " ", " ", " ", " ", " ", 7654321, "OTRO CLIENTE"]
    # ]

    # Convertir cada fila del DataFrame en lista y rellenar hasta 10 columnas
    nuevas_filas = []

    for _, row in df_total.iterrows():
        fila = [
            row[0],           # 0: TIPO (ej. "CAM-")
            row[1],          # 1: Código transportista
            row[2],          # 2: Código cliente
            " ",                   # 3
            " ",                   # 4
            " ",                   # 5
            " ",                   # 6
            " ",                   # 7
            row[3],          # 8: Repetir código cliente
            row[4]         # 9: Nombre cliente
        ]
        nuevas_filas.append(fila)

    # Insertar filas
    fila_inicio = ws.max_row + 1
    for i, fila in enumerate(nuevas_filas):
        tipo = str(fila[0]).strip()
        fill = next((color for prefijo, color in COLORS.items() if tipo.startswith(prefijo)), None)

        for j, valor in enumerate(fila):
            celda = ws.cell(row=fila_inicio + i, column=j + 1, value=valor)
            # Pintar solo si hay valor no vacío (ni espacios) y hay color asignado
            if fill and str(valor).strip():
                celda.fill = fill

    # Guardar sin perder estilos
    wb.save(PROCESSED_FILE)